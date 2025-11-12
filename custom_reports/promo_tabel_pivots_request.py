#%%imports
import MagnumDB
import os
import pandas as pd
import io
import logging

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter,column_index_from_string

from dotenv import load_dotenv

try:
    load_dotenv('.env')
except:
    pass

#declare db connections
db_promo_tabel = {
    'db_ip':os.environ['TABEL_IP'],
    'db_port':os.environ['TABEL_PORT'],
    'db_service':os.environ['TABEL_SERVICE'],
    'db_login':os.environ['TABEL_LOGIN'],
    'db_pass':os.environ['TABEL_PASS'] 
}

db_dwh = {
    'db_ip':os.environ['DWH_IP'],
    'db_port':os.environ['DWH_PORT'],
    'db_service':os.environ['DWH_SERVICE'],
    'db_login':os.environ['DWH_LOGIN'],
    'db_pass':os.environ['DWH_PASS'] 
}

#%%
def sexy_xlsx(table_1: pd.DataFrame, table_2: pd.DataFrame, iol):
    wb = Workbook()
    df1 = wb.create_sheet("Сводная по ТМ", 0)
    df2 = wb.create_sheet("Сводная по SKU", 1)

    del wb['Sheet']
    for r_1 in dataframe_to_rows(table_1, index=False, header=False):
        df1.append(r_1)
    for r_2 in dataframe_to_rows(table_2, index=False, header=False):                               
        df2.append(r_2)

    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 70

    def add_outline_border(sheet, start_row, start_col, end_row, end_col):
        """
        Разукрасить границы выбранного диапазона
        """
        borderTop = Border(top = Side(style = 'thick'))
        borderBottom = Border(bottom = Side(style = 'thick'))
        borderLeft = Border(left = Side(style = 'thick'))                
        borderRight = Border(right = Side(style = 'thick')) 
        borderTopLeft = Border(left = Side(style = 'thick'), top=Side(style='thick'))
        borderTopRight = Border(right = Side(style = 'thick'), top=Side(style='thick'))
        borderBottomLeft = Border(left = Side(style = 'thick'), bottom = Side(style = 'thick'))
        borderBottomRight = Border(right = Side(style = 'thick'), bottom = Side(style = 'thick'))
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                sheet.cell(start_row, col).border = borderTop
                sheet.cell(end_row, col).border = borderBottom
                sheet.cell(row, start_col).border = borderLeft
                sheet.cell(row, end_col).border = borderRight

        sheet.cell(start_row, start_col).border = borderTopLeft
        sheet.cell(start_row, end_col).border = borderTopRight
        sheet.cell(end_row, start_col).border = borderBottomLeft
        sheet.cell(end_row, end_col).border = borderBottomRight
        return sheet

    city_color_dict = {
        'Алматы':'ccffff',
        'Астана' : 'fcdbc0',
        'Караганда' : 'ccecff',
        'Шымкент' : 'ffff99',
        'Талдыкорган' : '92d050',
        'Петропавловск' : '71ffb1',
        'Кызылорда' : 'ff6915',
        'Тараз' : 'ffc1ff',
        'Туркестан' : 'ffe699',
        'Усть-Каменогорск' : '9999ff'
    }

    # ЛИСТ 1

    dfl_last_columns_letter = str(get_column_letter(df1.max_column))

    # Корректировка называний
    df1['A1'].value = ''
    df1['A2'].value = 'Подгруппа'
    df1['B1'].value = ''
    df1['B2'].value = 'Торговая марка'

    # Словарь с буквами городов
    city_column_dict = {city: [] for city in city_color_dict.keys()}

    # Форматы на первой строке
    for row in df1[f'A1' : dfl_last_columns_letter + '1']:  # Adjusted row range
        for cell in row:
            if cell.value in city_color_dict:
                city_column_dict[cell.value].append(cell.column_letter)
                cell.fill = PatternFill(start_color=city_color_dict[cell.value], end_color=city_color_dict[cell.value], fill_type='solid')
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.font = Font(bold=True)
                df1.column_dimensions[cell.column_letter].width = 15

    # Форматы на второй строке
    for row in df1[f'A2' : dfl_last_columns_letter + '2']:  # Adjusted row range
        for cell in row:
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # ФИЛЬТР НА ВСЮ ВТОРУЮ СТРОКУ
    df1.auto_filter.ref = f'A2:{str(get_column_letter(df1.max_column))}2'  # Apply filter to data range

    df1.column_dimensions['A'].width = 50
    df1.column_dimensions['B'].width = 20

    # Границы городов
    for cities, columns in city_column_dict.items():
        try:
            df1 = add_outline_border(df1, 1, column_index_from_string(columns[0]), df1.max_row, column_index_from_string(columns[-1]))
            df1 = add_outline_border(df1, 1, column_index_from_string(columns[0]), 1, column_index_from_string(columns[-1]))
            df1 = add_outline_border(df1, 2, column_index_from_string(columns[0]), 2, column_index_from_string(columns[-1]))
        except:
            pass
    
    # Границы основных названий
    for row in df1[f'A2' : 'B2']:  # Adjusted row range
        for cell in row:
            cell.border = Border(top=Side(border_style="thick"), left=Side(border_style="thick"), right=Side(border_style="thick"), bottom=Side(border_style="thick"))

    # Заморозить вторую строку
    df1.freeze_panes = df1[f'A3'] 

    # ЛИСТ 2

    df2_last_columns_letter = str(get_column_letter(df2.max_column))

    # Корректировка называний
    df2['A1'].value = ''
    df2['A2'].value = 'Подгруппа'
    df2['B1'].value = ''
    df2['B2'].value = 'Торговая марка'
    df2['C1'].value = ''
    df2['C2'].value = 'Код товара'
    df2['D1'].value = ''
    df2['D2'].value = 'Наименование товара'

    # Словарь с буквами городов
    city_column_dict = {city: [] for city in city_color_dict.keys()}

    # Форматы на первой строке
    for row in df2[f'A1' : df2_last_columns_letter + '1']:  # Adjusted row range
        for cell in row:
            if cell.value in city_color_dict:
                city_column_dict[cell.value].append(cell.column_letter)
                cell.fill = PatternFill(start_color=city_color_dict[cell.value], end_color=city_color_dict[cell.value], fill_type='solid')
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.font = Font(bold=True)
                df2.column_dimensions[cell.column_letter].width = 15

    # Форматы на второй строке
    for row in df2[f'A2' : df2_last_columns_letter + '2']:  # Adjusted row range
        for cell in row:
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # ФИЛЬТР НА ВСЮ ВТОРУЮ СТРОКУ
    df2.auto_filter.ref = f'A2:{str(get_column_letter(df2.max_column))}2'  # Apply filter to data range

    df2.column_dimensions['A'].width = 50
    df2.column_dimensions['B'].width = 20
    df2.column_dimensions['C'].width = 14
    df2.column_dimensions['D'].width = 50

    # Границы городов
    for cities, columns in city_column_dict.items():
        try:
            df2 = add_outline_border(df2, 1, column_index_from_string(columns[0]), df2.max_row, column_index_from_string(columns[-1]))
            df2 = add_outline_border(df2, 1, column_index_from_string(columns[0]), 1, column_index_from_string(columns[-1]))
            df2 = add_outline_border(df2, 2, column_index_from_string(columns[0]), 2, column_index_from_string(columns[-1]))
        except:
            pass

    # Границы основных названий
    for row in df2[f'A2' : 'D2']:  # Adjusted row range
        for cell in row:
            cell.border = Border(top=Side(border_style="thick"), left=Side(border_style="thick"), right=Side(border_style="thick"), bottom=Side(border_style="thick"))
                
    # Заморозить вторую строку
    df2.freeze_panes = df2[f'A3'] 

    return wb.save(iol)

# %%
def get_promo_tabel_pivots_data(promo_id, user_name, limit):

    if limit != None:
        lim_row = f'LIMIT {limit}'
    else:
        lim_row = ''

    # факт цен из табеля
    select_promos_TM_query  = open("custom_reports/promo_tabel_pivot_query_tabel.sql", encoding='utf-8', mode="r").read()
    db = MagnumDB.DBConnection(**db_promo_tabel)
    data_tabel = db.select(select_promos_TM_query%(promo_id, user_name, user_name, user_name, lim_row))
    db.close()


    trash_chars = {"'" : ""
                , "C" : "С" # eng -> rus
                , "T" : "Т" # eng -> rus
                , "НЕУЧАСТВУЕТ" : "НЕ УЧАСТВУЕТ"
                , "НЕУЧАСТСВУЕТ" : "НЕ УЧАСТВУЕТ"
                }

    main = data_tabel.copy()

    # Замена мусорных символов
    main['dmp_id_fix'] = main['dmp_id'].apply(lambda x: x.upper())

    for k, v in trash_chars.items():
        main['dmp_id_fix'] = main['dmp_id_fix'].str.replace(k, v)

    # Замена dmp_id на корректные
    main['dmp_id_fix'] = main['dmp_id_fix'].apply(lambda x: 'Т' if '-' in x else 
                                                    ('П' if x.isdigit() else x))

    # Скрыть пустые
    main = main[main['dmp_id_fix'] != ''].drop_duplicates().reset_index(drop=True)

    # Таблица на уровне ТМ
    # Сумирование дубликатов
    TM_pivot = main.groupby(['l5', 'brand', 'region_name', 'dmp_id_fix'])['dmp_id_count'].sum().reset_index()

    # Вывести регионы и ДМП в колонки
    TM_pivot = TM_pivot.pivot(index=[col for col in TM_pivot.columns if col not in ['region_name', 'dmp_id_count', 'dmp_id_fix']], columns=['region_name', 'dmp_id_fix'], values='dmp_id_count').reset_index()

    # Сортировка колонок
    TM_pivot = TM_pivot[[('l5', ''), ('brand', '')] + sorted([col for col in TM_pivot.columns if col not in [('l5', ''), ('brand', ''), ]])]

    # Сортировка and rename строк
    TM_pivot = TM_pivot.sort_values(['brand', 'l5']).reset_index(drop=True)

    TM_pivot = TM_pivot.rename(columns={'l5' : 'Подгруппа'
                                        ,'brand' : 'Торговая марка'})

    # Скрыть имена мультииндекса
    TM_pivot.columns.names = ['', '']

    # дропнуть нулевой уровень
    TM_pivot.loc[-1] = TM_pivot.columns.droplevel(0)
    TM_pivot.index = TM_pivot.index + 1

    # дропнуть первый уровень
    TM_pivot.loc[-1] = TM_pivot.columns.droplevel(1)
    TM_pivot.index = TM_pivot.index + 1

    # Сортировать индекс
    TM_pivot = TM_pivot.sort_index() 

    # Таблица на уровне SKU
    # Сумирование дубликатов
    SKU_pivot = main.groupby(['l5', 'brand', 'source_product_id', 'name_ware', 'region_name', 'dmp_id_fix'])['dmp_id_count'].sum().reset_index()

    # Вывести регионы и ДМП в колонки
    SKU_pivot = SKU_pivot.pivot(index=[col for col in SKU_pivot.columns if col not in ['region_name', 'dmp_id_count', 'dmp_id_fix']]
                                , columns=['region_name', 'dmp_id_fix']
                                , values='dmp_id_count').reset_index()

    # Сортировка колонок
    SKU_pivot = SKU_pivot[[('l5', ''), ('brand', ''), ('source_product_id', ''), ('name_ware', '')] + sorted([col for col in SKU_pivot.columns if col not in [('l5', ''), ('brand', ''), ('source_product_id', ''), ('name_ware', ''), ]])]

    # Сортировка and rename строк
    SKU_pivot = SKU_pivot.sort_values(['brand', 'l5', 'name_ware']).reset_index(drop=True)

    SKU_pivot = SKU_pivot.rename(columns={'l5' : 'Подгруппа'
                                        ,'brand' : 'Торговая марка'
                                        ,'source_product_id' : 'Код товара'
                                        ,'name_ware' : 'Наименование товара'})

    # Скрыть имена мультииндекса
    SKU_pivot.columns.names = ['', '']

    # дропнуть нулевой уровень
    SKU_pivot.loc[-1] = SKU_pivot.columns.droplevel(0)
    SKU_pivot.index = SKU_pivot.index + 1

    # дропнуть первый уровень
    SKU_pivot.loc[-1] = SKU_pivot.columns.droplevel(1)
    SKU_pivot.index = SKU_pivot.index + 1

    # Сортировать индекс
    SKU_pivot = SKU_pivot.sort_index()

    # to binary flow
    output = io.BytesIO()
    sexy_xlsx(TM_pivot, SKU_pivot, output)


    return io.BytesIO(output.getvalue())