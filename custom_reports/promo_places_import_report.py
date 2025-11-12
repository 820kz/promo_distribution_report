#%%
import MagnumDB
import os
import pandas as pd
import io
import re
import numpy
import logging

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


pd.set_option("display.max_columns", None)
pd.options.display.float_format = '{:,}'.format

#%%read params
from dotenv import load_dotenv

try:
    load_dotenv('.env')
except:
    pass

def sexy_xlsx(table: pd.DataFrame, iol):
    wb = Workbook()
    df3 = wb.create_sheet("ДМП", 0)
    del wb['Sheet']

    # Add 6 empty rows at the beginning
    for _ in range(0):
        df3.append([])

    for r_3 in dataframe_to_rows(table, index=True, header=True):                               
        df3.append(r_3)
        
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 70

    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=10)

    ws.column_dimensions = dim_holder
        
    thin = Side(border_style="thin", color="000000")

    df3.delete_cols(1)

    header_row = 1 # строка с заголовком
    black_row = header_row + 1 # черная полоса с фильтром
    first_values_row = header_row + 1 # первая строка со значениями
    last_values_row = header_row + 1 + len(table) # последняя строка со значениями

    right_bottom_cell = str(get_column_letter(ws.max_column)) + str(ws.max_row) # правый нижний угол таблицы
    top_right_black_row = str(get_column_letter(ws.max_column))+str(black_row) # правый край черной полоски

    # ЗАГОЛОВКИ КОЛОНОК
    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:  # Adjusted row range
        for cell in row:
            cell.fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin) 
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # ФИЛЬТР НА ЧЕРНОЙ ПОЛОСЕ
    df3.auto_filter.ref = f'A{black_row}:{right_bottom_cell}'  # Apply filter to data range

    def borders_correction(start, end, left, right, top, bottom):
        """
        Подкрасить границы диапазона городов
        """
        df3[start + str(end)].border = Border(left=Side(style=left), 
                                        right=Side(style=right), 
                                        top=Side(style=top), 
                                        bottom=Side(style=bottom))
        
    def shops_bordering(cities, color, shop_filter, target_columns):
        try:
            # рисовать границы на магазинах или в кастомном месте
            if shop_filter == True:
                # залить заголовки городов разными цветами
                shops_cells = []
                for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:  # Adjusted row range
                    for cell in row:
                        # если в заголовке больше двух слов
                        if len(cell.value.split()) > 1:
                            # если второе слово равно городу...
                            if cell.value.split()[1] in cities:
                                shops_cells.append(cell)
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        else:
                            # если второе слово равно городу...
                            if cell.value in cities:
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

                start_column = shops_cells[0].column_letter
                end_column = shops_cells[-1].column_letter
            else:
                shops_cells = target_columns
                start_column = shops_cells[0]
                end_column = shops_cells[-1]
        except:
            # если нет какого-то города
            pass

    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:  # Adjusted row range
        for cell in row:
            if cell.value == 'ТЗ\nТФ\n№3\nEXPRESS':
                cell.value = 'ТЗ\nШФ\n№999\nEXPRESS'

    # Кинуть границы динамически на все магазины
    shops_bordering(cities = ['АФ', 'ЕКФ','КПФ','ФКС'], color = 'ccffff', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['АСТ'], color = 'fcdbc0', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['КФ'], color = 'ccecff', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ШФ'], color = 'ffff99', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ТКФ'], color = '92d050', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ППФ'], color = '71ffb1', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['КЗФ'], color = 'ff6915', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ТЗФ'], color = 'ffc1ff', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ТФ'], color = 'ffe699', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['УКФ'], color = '9999ff', shop_filter = True, target_columns = None)

    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:
        for cell in row:
            if cell.value == 'Акционная цена с НДС, тг':
                cell.number_format = '_-* # ##0_-;-* # ##0_-;_-* "-"??_-;_-@_-'
                cell.fill = PatternFill(start_color='FF5D5D', end_color='FF5D5D', fill_type='solid')
                df3.column_dimensions[cell.column_letter].width = 14
                for ccell in range(0, last_values_row - header_row + 1):
                    if ccell < last_values_row - header_row:
                        df3[f'{cell.column_letter}{first_values_row+ccell}'].number_format = '_-* # ##0_-;-* # ##0_-;_-* "-"??_-;_-@_-'
                #cell.width = 14 

    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:
        for cell in row:
            if cell.value == 'Скидка для покупателя, %':
                cell.number_format = '0.0%'
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                df3.column_dimensions[cell.column_letter].width = 14
                for ccell in range(0, last_values_row - header_row + 1):
                    if ccell < last_values_row - header_row:
                        df3[f'{cell.column_letter}{first_values_row+ccell}'].number_format = '0.0%'
                
    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:
        for cell in row:
            if cell.value == 'Регулярная розничная цена с НДС,тг':
                cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
                df3.column_dimensions[cell.column_letter].width = 14
                for ccell in range(0, last_values_row - header_row + 1):
                    if ccell < last_values_row - header_row:
                        df3[f'{cell.column_letter}{first_values_row+ccell}'].number_format = '_-* # ##0_-;-* # ##0_-;_-* "-"??_-;_-@_-'

    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:
        for cell in row:
            if cell.value == 'Наименование поставщика':
                df3.column_dimensions[cell.column_letter].width = 40
            elif cell.value == 'Наименование товара':
                df3.column_dimensions[cell.column_letter].width = 75
            elif cell.value == 'Код товара в Спрут':
                df3.column_dimensions[cell.column_letter].width = 13
            elif cell.value == 'Наименование товара':
                df3.column_dimensions[cell.column_letter].width = 75
            elif cell.value == 'Штрих-код':
                df3.column_dimensions[cell.column_letter].width = 17
            elif cell.value == 'Подгруппа':
                df3.column_dimensions[cell.column_letter].width = 40
            elif cell.value == 'Группа':
                df3.column_dimensions[cell.column_letter].width = 20
            elif cell.value == 'Отдел':
                df3.column_dimensions[cell.column_letter].width = 20
            elif cell.value in ['Торговая марка', 'Тип ДМП']:
                df3.column_dimensions[cell.column_letter].width = 14

    # ТФ3 должен стоять рядом с ШФ
    # переименовываем его обратно в ТФ3 и красим в цвет ТФ
    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:  # Adjusted row range
        for cell in row:
            if cell.value == 'ТЗ\nШФ\n№999\nEXPRESS':
                cell.value = 'ТЗ\nТФ\n№3\nEXPRESS'
                cell.fill = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')


    # ЧЕРНАЯ ЛИНИЯ ПОД ЗАГОЛОВКАМИ С ФИЛЬТРОМ
    for row in df3[f'A{black_row}':top_right_black_row]:  # Adjusted row range
        for cell in row:
            cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    #df3.column_dimensions['Q'].hidden= True

    df3.freeze_panes = df3[f'A{str(first_values_row)}'] 

    return  wb.save(iol)

def sort_key(header):
    '''
    Ручной сортировщик колонок
    '''
    match = re.search(r'(\D+)(\d+)', header)
    if match:
        prefix, number = match.groups()
        if prefix.split()[1] == 'АФ':
            prefix = 'AA'
        elif prefix.split()[1] == 'ФКС':
            prefix = 'AB'
        elif prefix.split()[1] == 'ЕКФ':
            prefix = 'AC'
        elif prefix.split()[1] == 'КПФ':
            prefix = 'AD'
        elif prefix.split()[1] == 'АСТ':
            prefix = 'AE'
        elif prefix.split()[1] == 'КФ':
            prefix = 'AF'
        elif prefix.split()[1] == 'ШФ':
            prefix = 'AGA'
        elif prefix.split()[1] == 'ТКФ':
            prefix = 'AH'
        elif prefix.split()[1] == 'ППФ':
            prefix = 'AI'
        elif prefix.split()[1] == 'КЗФ':
            prefix = 'AJ'
        elif prefix.split()[1] == 'ТЗФ':
            prefix = 'AK'
        elif prefix.split()[1] == 'ТФ':
            # ТФ3 должен стоять рядом с Шымкентом
            if number == '3':
                prefix = 'AGB'
            else:
                prefix = 'AL'
        elif prefix.split()[1] == 'УКФ':
            prefix = 'AM'
        return (prefix, int(number))
    else:
        return (header, 0)
    
#declare db connections
db_promo_tabel = {
    'db_ip':os.environ['TABEL_IP'],
    'db_port':os.environ['TABEL_PORT'],
    'db_service':os.environ['TABEL_SERVICE'],
    'db_login':os.environ['TABEL_LOGIN'],
    'db_pass':os.environ['TABEL_PASS'] 
}

# %%
def get_promo_places_import_data(promo_id, user_name, limit):

    if limit != None:
        lim_row = f'LIMIT {limit}'
    else:
        lim_row = ''

    # Все промо-места из текущего promo_id
    select_tabel_query  = open("custom_reports/promo_places_import_query_tabel_main.sql", encoding='utf-8', mode="r").read()
    db = MagnumDB.DBConnection(**db_promo_tabel)
    data = db.select(select_tabel_query%(promo_id, user_name, user_name, user_name, lim_row))
    db.close()

    data['shop_name'] = data['shop_name'].str.replace(' ', '\n')
    data['shop_name'] = data['shop_name'].str.replace('ТЗ\nТФ\n№3\nEXPRESS', 'ТЗ\nШФ\n№999\nEXPRESS')

    tabel_data = data.replace({numpy.nan: None}).copy().drop(columns=['Регулярная розничная цена с НДС,тг',
        'Акционная цена с НДС, тг', 'Скидка для покупателя, %', 'Тип ДМП'])
    #tabel_data = data.replace({numpy.nan: None}).copy()

    main_cols_data = tabel_data[['Отдел', 'Группа', 'Подгруппа', 'Штрих-код',
    'Код товара в Спрут', 'Наименование товара', 'Торговая марка']].drop_duplicates().reset_index(drop=True)

    regions=['Алматы',
            'Астана',
            'Караганда',
            'Шымкент',
            'Талдыкорган',
            'Петропавловск',
            'Кызылорда',
            'Тараз',
            'Туркестан',
            'Усть-Каменогорск']

    for region in regions:

        barcode_data = tabel_data[(tabel_data['Регион проведения акции'] == region)].drop(columns='Регион проведения акции').copy()
        barcode_data = barcode_data[['Код товара в Спрут', 'Штрих-код'] + [col for col in barcode_data.columns if col not in ['Отдел', 'Группа', 'Подгруппа', 'Штрих-код',
            'Код товара в Спрут', 'Наименование товара', 'Торговая марка']]]

        if barcode_data.size > 0:
            try:
                barcode_pivot = barcode_data.pivot(
                                    index=[col for col in barcode_data.columns if col not in ['shop_name', 'dmp_id']], 
                                    columns='shop_name', 
                                    values='dmp_id'
                                ).reset_index()
                barcode_pivot = barcode_pivot[[col for col in barcode_data.columns if col not in ['shop_name', 'dmp_id']] + sorted(barcode_data['shop_name'].unique(), key=sort_key)]
            except:
                barcode_pivot['shop_name'] = barcode_pivot['shop_name'].str.replace('ТЗ\nШФ\n№999\nEXPRESS', 'ТЗ\nТФ\n№3\nEXPRESS')
                if len(barcode_data[barcode_data.duplicated()]) > 0:
                    logging.error(f"Дубликаты {barcode_data[barcode_data.duplicated()]['Штрих-код'].unique()} в {barcode_data[barcode_data.duplicated()]['shop_name'].unique()}")
                else:
                    barcode_data = barcode_data[[col for col in barcode_data.columns if col not in ['dmp_id']]]
                    logging.error(f"Дубликаты {barcode_data[barcode_data.duplicated()]['Штрих-код'].unique()} в {barcode_data[barcode_data.duplicated()]['shop_name'].unique()}")


            main_cols_data = pd.merge(main_cols_data, barcode_pivot, on=['Код товара в Спрут', 'Штрих-код'], how='left', suffixes=(f'|{region}', f'|{region}'))

    main_cols_data.columns = [col.split('|')[0] for col in main_cols_data.columns]

    # to binary flow
    output = io.BytesIO()
    sexy_xlsx(main_cols_data, output)

    return io.BytesIO(output.getvalue())


