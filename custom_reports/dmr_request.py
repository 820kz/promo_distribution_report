#%%imports
import MagnumDB
import os
import pandas as pd
import io
import time
import re
from numpy import where
from itertools import chain
import logging

pd.set_option("display.max_columns", None)
pd.options.display.float_format = '{:,}'.format

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

#%%read params


from dotenv import load_dotenv

try:
    load_dotenv('.env')
except:
    pass

#declare db connections
db_promo_tabel = {
    'db_ip':os.environ['TABEL_IP'],
    'db_port':os.environ['TABEL_PORT'],
    'db_service':os.environ['TABEL_SERVICE'],
    'db_login':os.environ['TABEL_LOGIN'],
    'db_pass':os.environ['TABEL_PASS'] 
}

colms = {'РН' : 'A',
'КМ' : 'B',
'Отдел' : 'C',
'Группа' : 'D',
'Подгруппа' : 'E',
'АФ' : 'F',
'АСТ' : 'G',
'КФ' : 'H',
'ШФ' : 'I',
'ТКФ' : 'J',
'ППФ' : 'K',
'КЗФ' : 'L',
'ТЗФ' : 'M',
'ТФ' : 'N',
'УКФ' : 'O',
'Регион проведения акции' : 'P',
'Форматы' : 'Q',
'Топ товаров' : 'R',
'Модуль' : 'S',
'Штрих-код' : 'T',
'Код товара в Спрут' : 'U',
'Наименование товара' : 'V',
'Наименование поставщика' : 'W',
'Торговая марка' : 'X',
'Регулярная розничная цена с НДС,тг' : 'Y',
'Акционная цена с НДС, тг' : 'Z',
'Скидка для покупателя, %' : 'AA',
'Исключить ТК Алматы' : 'AB',
'Исключить ТК Астаны' : 'AC',
'Исключить ТК Караганды' : 'AD',
'Исключить ТК Шымкент' : 'AE',
'Исключить ТК Талдыкорган' : 'AF',
'Исключить ТК Петропавловск' : 'AG',
'Исключить ТК Кызылорда' : 'AH',
'Исключить ТК Тараз' : 'AI',
'Исключить ТК Туркестан' : 'AJ',
'Исключить ТК Усть-Каменогорск' : 'AK'}


def sexy_xlsx(table: pd.DataFrame, iol):
    wb = Workbook()
    df3 = wb.create_sheet("ДМП", 0)
    del wb['Sheet']
    
    # Add 6 empty rows at the beginning
    for _ in range(4):
        df3.append([])

    for r_3 in dataframe_to_rows(table, index=True, header=True):                               
        df3.append(r_3)
        
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 70

    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=10)

    ws.column_dimensions = dim_holder
        
    thin = Side(border_style="thin", color="000000")

    df3.delete_cols(1)

    header_row = 5 # строка с заголовком
    black_row = header_row + 1 # черная полоса с фильтром
    first_values_row = header_row + 2 # первая строка со значениями
    last_values_row = header_row + 1 + len(table) # последняя строка со значениями

    right_bottom_cell = str(get_column_letter(ws.max_column)) + str(ws.max_row) # правый нижний угол таблицы
    top_right_black_row = str(get_column_letter(ws.max_column))+str(black_row) # правый край черной полоски
    
    # ЗАГОЛОВКИ КОЛОНОК
    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:  # Adjusted row range
        for cell in row:
            cell.fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin) 
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # СЧЕТЧИК ГОРОДОВ ПО МОДУЛЯМ - ЗАГОЛОВОК
    for row in df3['F1':'O1']:  # Adjusted row range
        for cell in row:
            cell.fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin) 
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # СЧЕТЧИК ГОРОДОВ ПО МОДУЛЯМ - ЗНАЧЕНИЯ
    for row in df3['F2':'O2']:  # Adjusted row range
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin) 
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # ФИЛЬТР НА ЧЕРНОЙ ПОЛОСЕ
    df3.auto_filter.ref = f'A{black_row}:{right_bottom_cell}'  # Apply filter to data range

    # values = ["Алматы", "Астана", "Караганда", "Кызылорда", "Петропавловск", "Талдыкорган", "Тараз", "Туркестан", "Усть-Каменогорск", "Шымкент"]
    # for col_idx, value in enumerate(values, start=2):  # Start from column B (index 2)
    #     df3.cell(row=1, column=col_idx, value=value)

    # НАИМЕНОВАНИЯ МОДУЛЕЙ НАД ОСНОВНОЙ ТАБЛИЦЕЙ
    df3['F1'] = 'Алматы'
    df3['G1'] = 'Астана'
    df3['H1'] = 'Караганда'
    df3['I1'] = 'Шымкент'
    df3['J1'] = 'Талдыкорган'
    df3['K1'] = 'Петропавловск'
    df3['L1'] = 'Кызылорда'
    df3['M1'] = 'Тараз'
    df3['N1'] = 'Туркестан'
    df3['O1'] = 'Усть-Каменогорск'    

    # ФОРМУЛА МОДУЛЕЙ НАД ОСНОВНОЙ ТАБЛИЦЕЙ
    df3['F2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&F1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'
    df3['G2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&G1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'
    df3['H2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&H1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'
    df3['I2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&I1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'
    df3['J2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&J1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'
    df3['K2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&K1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'
    df3['L2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&L1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'
    df3['M2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&M1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'
    df3['N2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&N1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'
    df3['O2'] = f'=COUNTIFS(${colms["Регион проведения акции"]}${first_values_row}:${colms["Регион проведения акции"]}${last_values_row},"*"&O1&"*",${colms["Модуль"]}${first_values_row}:${colms["Модуль"]}${last_values_row},1)'

    df3.column_dimensions[colms['РН']].width = 20
    df3.column_dimensions[colms['КМ']].width = 20
    df3.column_dimensions[colms['Отдел']].width = 20
    df3.column_dimensions[colms['Группа']].width = 20
    df3.column_dimensions[colms['Подгруппа']].width = 40

    df3.column_dimensions[colms['Регион проведения акции']].width = 18
    df3.column_dimensions[colms['Топ товаров']].width = 18
    df3.column_dimensions[colms['Форматы']].width = 28
    df3.column_dimensions[colms['Модуль']].width = 13
    df3.column_dimensions[colms['Штрих-код']].width = 17
    df3.column_dimensions[colms['Код товара в Спрут']].width = 13
    df3.column_dimensions[colms['Наименование товара']].width = 75
    df3.column_dimensions[colms['Наименование поставщика']].width = 40

    # Увеличение ширины колонок с Z до AN на 12
    for col in [colms['Торговая марка'],
                 colms['Регулярная розничная цена с НДС,тг'], 
                 colms['Акционная цена с НДС, тг'], 
                 colms['Скидка для покупателя, %'], 
                 colms['Исключить ТК Алматы'], colms['Исключить ТК Астаны'], colms['Исключить ТК Караганды'], 
                 colms['Исключить ТК Шымкент'], colms['Исключить ТК Талдыкорган'], colms['Исключить ТК Петропавловск'], 
                 colms['Исключить ТК Кызылорда'], colms['Исключить ТК Тараз'], colms['Исключить ТК Туркестан'], 
                 colms['Исключить ТК Усть-Каменогорск']]:
        df3.column_dimensions[col].width = 14 

    # # залить все ячейки пунктирной границей
    # for cell in range(2, last_values_row - header_row + 1): 
    #     for row in df3[f'A{header_row+cell}':get_column_letter(ws.max_column)+f"{header_row+cell}"]:  # Adjusted row range
    #         for cell in row:
    #             cell.border = Border(left=Side(style='dashed'), 
    #                                 right=Side(style='dashed'), 
    #                                 top=Side(style='dashed'), 
    #                                 bottom=Side(style='dashed'))


    def borders_correction(start, end, left, right, top, bottom):
        """
        Подкрасить границы диапазона городов
        """
        df3[start + str(end)].border = Border(left=Side(style=left), 
                                        right=Side(style=right), 
                                        top=Side(style=top), 
                                        bottom=Side(style=bottom))
        
    def shops_bordering(cities, color, shop_filter, target_columns):
        try:
            # рисовать границы на магазинах или в кастомном месте
            if shop_filter == True:
                # залить заголовки городов разными цветами
                shops_cells = []
                for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:  # Adjusted row range
                    for cell in row:
                        # если в заголовке больше двух слов
                        if len(cell.value.split()) > 1:
                            # если второе слово равно городу...
                            if cell.value.split()[1] in cities:
                                shops_cells.append(cell)
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        else:
                            # если второе слово равно городу...
                            if cell.value in cities:
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

                start_column = shops_cells[0].column_letter
                end_column = shops_cells[-1].column_letter
            else:
                shops_cells = target_columns
                start_column = shops_cells[0]
                end_column = shops_cells[-1]

            # залить первую левую границу колонки АФ и последнюю правую АФ
            for cell in range(0, last_values_row - header_row + 1): 
                df3[start_column + f'{header_row+cell}'].border = Border(left=Side(style='thick'))
                
                df3[end_column + f'{header_row+cell}'].border = Border(right=Side(style='thick'))
                
            # Залить верхнюю гарницу магазинов АФ жирным
            for row in df3[start_column+str(header_row) : end_column+str(header_row)]:
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), 
                                            right=Side(style='thin'), 
                                            top=Side(style='thick'), 
                                            bottom=Side(style='thin'))

            # Залить нижнюю гарницу магазинов АФ жирным
            for row in df3[start_column + str(last_values_row) : end_column + str(last_values_row)]:
                for cell in row:
                    cell.border = Border(bottom=Side(style='thick'))
            
            # Докрасить левый верхний угол АФ
            borders_correction(start = start_column, end = header_row, left = 'thick', right = 'thin', top = 'thick', bottom = 'thin')
            # Докрасить правый верхний угол АФ
            borders_correction(start = end_column, end = header_row, left = 'thin', right = 'thick', top = 'thick', bottom = 'thin')
            # Докрасить левый нижний угол АФ
            borders_correction(start = start_column, end = last_values_row, left = 'thick', right = None, top = None, bottom = 'thick')
            # Докрасить правый нижний угол АФ
            borders_correction(start = end_column, end = last_values_row, left = None, right = 'thick', top = None, bottom = 'thick')
        except:
            # если нет какого-то города
            pass

    # Кинуть границы на всю таблицу
    #shops_bordering(cities = None, color = None, shop_filter = False, target_columns = ['A', str(get_column_letter(ws.max_column))])

    # ТФ3 должен стоять рядом с ШФ
    # переименовываем его в ШФ
    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:  # Adjusted row range
        for cell in row:
            if cell.value == 'ТЗ\nТФ\n№3\nEXPRESS':
                cell.value = 'ТЗ\nШФ\n№999\nEXPRESS'

    # Кинуть границы динамически на все магазины
    shops_bordering(cities = ['АФ', 'ЕКФ','КПФ','ФКС'], color = 'ccffff', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['АСТ'], color = 'fcdbc0', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['КФ'], color = 'ccecff', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ШФ'], color = 'ffff99', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ТКФ'], color = '92d050', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ППФ'], color = '71ffb1', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['КЗФ'], color = 'ff6915', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ТЗФ'], color = 'ffc1ff', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['ТФ'], color = 'ffe699', shop_filter = True, target_columns = None)
    shops_bordering(cities = ['УКФ'], color = '9999ff', shop_filter = True, target_columns = None)

    # ТФ3 должен стоять рядом с ШФ
    # переименовываем его обратно в ТФ3 и красим в цвет ТФ
    for row in df3[f'A{header_row}':get_column_letter(ws.max_column)+f"{header_row}"]:  # Adjusted row range
        for cell in row:
            if cell.value == 'ТЗ\nШФ\n№999\nEXPRESS':
                cell.value = 'ТЗ\nТФ\n№3\nEXPRESS'
                cell.fill = PatternFill(start_color='ffe699', end_color='ffe699', fill_type='solid')

    # Колонки исключить
    shops_bordering(cities = None, color = None, shop_filter = False, target_columns = [colms['Исключить ТК Алматы'], colms['Исключить ТК Усть-Каменогорск']])

    # Еще колонки с магазинами
    shops_bordering(cities = None, color = None, shop_filter = False, target_columns = [colms['АФ'], colms['УКФ']])

    # ФОРМУЛЫ В ОСНОВНОЙ ТАБЛИЦЕ

    # Подровнять колонки с городами
    for cell in range(0, last_values_row - header_row - 1): 
        for cc in [colms['АФ'], colms['АСТ'], colms['КФ'], colms['ШФ'], colms['ТКФ'], 
                   colms['ППФ'], colms['КЗФ'], colms['ТЗФ'], colms['ТФ'], colms['УКФ']]:
            df3[f'{cc}{first_values_row+cell}'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            df3[f'{cc}{first_values_row+cell}'].font = Font(bold=True)

    # Модули
    for cell in range(0, last_values_row - header_row - 1):
        if df3[f'{colms["Скидка для покупателя, %"]}{first_values_row+cell}'].value == 1:
            df3[f'{colms["Модуль"]}{first_values_row+cell}'].fill = PatternFill(start_color='f2b0b0', end_color='f2b0b0', fill_type='solid')

    # Регулярная розничная цена с НДС, тенге
    for cell in range(0, last_values_row - header_row + 1):
        if cell < last_values_row - header_row - 1:
            df3[f'{colms["Регулярная розничная цена с НДС,тг"]}{first_values_row+cell}'].number_format = '_-* # ##0_-;-* # ##0_-;_-* "-"??_-;_-@_-'
            df3[f'{colms["Регулярная розничная цена с НДС,тг"]}{first_values_row+cell}'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        df3[f'{colms["Регулярная розничная цена с НДС,тг"]}{first_values_row - 2 + cell}'].fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

    # Акционная цена с НДС, тенге
    for cell in range(0, last_values_row - header_row + 1):
        if cell < last_values_row - header_row - 1:
            df3[f'{colms["Акционная цена с НДС, тг"]}{first_values_row+cell}'].number_format = '_-* # ##0_-;-* # ##0_-;_-* "-"??_-;_-@_-'
            df3[f'{colms["Акционная цена с НДС, тг"]}{first_values_row+cell}'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        df3[f'{colms["Акционная цена с НДС, тг"]}{first_values_row - 2 + cell}'].fill = PatternFill(start_color='FF5D5D', end_color='FF5D5D', fill_type='solid')

    # Скидка для покупателя, %
    for cell in range(0, last_values_row - header_row + 1):
        if cell < last_values_row - header_row - 1:
            df3[f'{colms["Скидка для покупателя, %"]}{first_values_row+cell}'].number_format = '0%'
            df3[f'{colms["Скидка для покупателя, %"]}{first_values_row+cell}'] = f'=IFERROR(1-({colms["Акционная цена с НДС, тг"]}{first_values_row+cell}/{colms["Регулярная розничная цена с НДС,тг"]}{first_values_row+cell}),0)'
            df3[f'{colms["Скидка для покупателя, %"]}{first_values_row+cell}'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        df3[f'{colms["Скидка для покупателя, %"]}{first_values_row - 2 + cell}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # ЧЕРНАЯ ЛИНИЯ ПОД ЗАГОЛОВКАМИ С ФИЛЬТРОМ
    for row in df3[f'A{black_row}':top_right_black_row]:  # Adjusted row range
        for cell in row:
            cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    #df3.column_dimensions['Q'].hidden= True

    df3.freeze_panes = df3[f'A{str(first_values_row)}'] 

    return wb.save(iol)


def get_dmr_data(promo_id, user_name, limit):

    if limit != None:
        lim_row = f'LIMIT {limit}'
    else:
        lim_row = ''

    select_promos_query  = open("custom_reports/dmr_query.sql", encoding='utf-8', mode="r").read()

    db = MagnumDB.DBConnection(**db_promo_tabel)
    data = db.select(select_promos_query%(promo_id, promo_id, user_name, user_name, user_name, lim_row))
    db.close()

    db = MagnumDB.DBConnection(**db_promo_tabel)
    formats = db.select("select short_shop_name, region_name, shop_format from nsi.sprut_region_shops where short_shop_name like 'ТЗ%'")
    db.close()

    if 'DAILY' not in data['Форматы'].unique()[0].split(', '):
        formats = formats.loc[formats['shop_format']=='DAILY'].reset_index(drop=True)
    else:
        formats = formats.loc[formats['shop_format']!='DAILY'].reset_index(drop=True)

    # Колонки для уравнивания тф3 и шымкента
    excl_list = ['Регион проведения акции',
                    'Топ товаров', 
                    'Форматы', 
                    'Тип акции', 
                    'Модуль', 
                    'Наименование поставщика',
                    'Торговая марка', 
                    'Регулярная розничная цена с НДС,тг',
                    'Акционная цена с НДС, тг', 
                    'Скидка для покупателя, %']
    
    def update_sales_forecast_and_supplier(df):
        """
        Показатели для ТФ3 как для Шымкента
        """
        express_indices = df[df['shop_name'] == 'ТЗ ТФ №3 EXPRESS'].index
        for i in express_indices:
            sprut_code = df.loc[i, 'Код товара в Спрут']
            try:
                shymkent_index = df[(df['Код товара в Спрут'] == sprut_code) & (df['Регион проведения акции'] == 'Шымкент')].index[0]
                for cx in excl_list:
                    try:
                        df.loc[i, cx] = df.loc[shymkent_index, cx]
                    except:
                        # странная ошибка, но всё работает
                        pass
            except:
                # пока временно
                pass
    try:
        update_sales_forecast_and_supplier(data)
    except:
        # временно
        pass

    #data['Регион проведения акции'] = where((data['shop_name']=='ТЗ ТФ №3 EXPRESS')&(data['dmp_id']), 'Шымкент' , data['Регион проведения акции'])

    data['Регион проведения акции'] = where(data['shop_name'] == 'ТЗ ТФ №3 EXPRESS', 'Шымкент' , data['Регион проведения акции'])

    # названия магазов с новой строки
    data['shop_name'] = data['shop_name'].str.replace(' ', '\n')

    data['shop_name'] = data['shop_name'].str.replace('ТЗ\nТФ\n№3\nEXPRESS', 'ТЗ\nШФ\n№999\nEXPRESS')

    # agg dmp_id
    #pivoted_df = data.pivot(index=[col for col in data.columns if col not in ['shop_name', 'dmp_id']], columns='shop_name', values='dmp_id').reset_index()
    data = data[(data['dmp_id'].notna())&(data['dmp_id']!='')].reset_index(drop=True)
    try:
        pivoted_df = data.pivot(index=[col for col in data.columns if col not in ['shop_name', 'dmp_id']], columns='shop_name', values='dmp_id').reset_index()
    except:
        data['shop_name'] = data['shop_name'].str.replace('ТЗ\nШФ\n№999\nEXPRESS', 'ТЗ\nТФ\n№3\nEXPRESS')
        if len(data[data.duplicated()]) > 0:
            logging.error(f"Дубликаты {data[data.duplicated()]['Штрих-код'].unique()} в {data[data.duplicated()]['shop_name'].unique()}")
        else:
            pivoted_df = data[[col for col in data.columns if col not in ['dmp_id']]]
            logging.error(f"Дубликаты {pivoted_df[pivoted_df.duplicated()]['Штрих-код'].unique()} в {pivoted_df[pivoted_df.duplicated()]['shop_name'].unique()}")
    # # fill the empty
    # for shp in list(data['shop_name'].unique()):
    #     pivoted_df[shp] =  pivoted_df[shp].fillna('НЕ УЧАСТВУЕТ')#.replace(['НЕ УЧАСТВУЕТ', 'не участвует', 'Не участвует'], nan).fillna(nan)\

    def sort_key(header):
        '''
        Ручной сортировщик колонок
        '''
        match = re.search(r'(\D+)(\d+)', header)
        if match:
            prefix, number = match.groups()
            if prefix.split()[1] == 'АФ':
                prefix = 'AA'
            elif prefix.split()[1] == 'ФКС':
                prefix = 'AB'
            elif prefix.split()[1] == 'ЕКФ':
                prefix = 'AC'
            elif prefix.split()[1] == 'КПФ':
                prefix = 'AD'
            elif prefix.split()[1] == 'АСТ':
                prefix = 'AE'
            elif prefix.split()[1] == 'КФ':
                prefix = 'AF'
            elif prefix.split()[1] == 'ШФ':
                prefix = 'AGA'
            elif prefix.split()[1] == 'ТКФ':
                prefix = 'AH'
            elif prefix.split()[1] == 'ППФ':
                prefix = 'AI'
            elif prefix.split()[1] == 'КЗФ':
                prefix = 'AJ'
            elif prefix.split()[1] == 'ТЗФ':
                prefix = 'AK'
            elif prefix.split()[1] == 'ТФ':
                # ТФ3 должен стоять рядом с Шымкентом
                if number == '3':
                    prefix = 'AGB'
                else:
                    prefix = 'AL'
            elif prefix.split()[1] == 'УКФ':
                prefix = 'AM'
            return (prefix, int(number))
        else:
            return (header, 0)
        
    pivoted_df = pivoted_df[[col for col in data.columns if col not in ['shop_name', 'dmp_id']] + sorted(data['shop_name'].unique(), key=sort_key)]

    exclude_dict = {
        'АФ': 'Исключить ТК Алматы',
        'ФКС': 'Исключить ТК Алматы',
        'ЕКФ': 'Исключить ТК Алматы',
        'АСТ': 'Исключить ТК Астаны',
        'УКФ': 'Исключить ТК Усть-Каменогорск',
        'ШФ': 'Исключить ТК Шымкент',
        'ТКФ': 'Исключить ТК Талдыкорган',
        'ППФ': 'Исключить ТК Петропавловск',
        'КЗФ': 'Исключить ТК Кызылорда',
        'ТЗФ': 'Исключить ТК Тараз',
        'ТФ': 'Исключить ТК Туркестан',
        'КФ': 'Исключить ТК Караганды'
    }

    reverse_exclude_dict = {}
    for key, value in exclude_dict.items():
        reverse_exclude_dict.setdefault(value, []).append(key)

    #
    formats['short'] = formats['short_shop_name'].apply(lambda x: x.split(' ')[1])
    formats['short_shop_name'] = formats['short_shop_name'].str.replace('ТЗ ', '').str.replace(' №', '')
    ff = formats.groupby('short')['short_shop_name'].apply(list).to_dict()

    # exclude shops columns
    def fill_exclude_columns(row, columns):
        """
        Заполнить колонки "Исключить ТК ..." 
        названиями колонок с ДМП - НЕ УЧАСТВУЕТ
        """
        exclude_columns = {category: [] for category in exclude_dict.values()}

        for col in columns:
            for keyword, target_column in exclude_dict.items():
                if keyword in col and ('НЕ УЧАСТВУЕТ' in str(row[col])):
                    clean_col = col.replace('ТЗ\n', '').replace('\n№', '').replace('\nEXPRESS', '').replace('\nSUPER', '').replace('\nDAILY', '').replace('\nHYPER', '')
                    if clean_col == 'ТФ3':
                        exclude_columns['Исключить ТК Шымкент'].append(clean_col)
                    else:
                        words = col.split()
                        if words[1] == keyword:
                            exclude_columns[target_column].append(clean_col)

        for target_column, exclude_list in exclude_columns.items():
            additional_shops = list(chain(*[i for i in [ff.get(key) for key in reverse_exclude_dict[target_column]] if i is not None]))
            row[target_column] = ', '.join(exclude_list + (additional_shops if row[[col for col in columns if '\n' in col and col.split('\n')[1] in reverse_exclude_dict[target_column]]].notnull().any() else []))
            if pd.api.types.is_numeric_dtype(row[target_column]):
                row[target_column] = row[target_column].astype(str)
        return row

    pivoted_df = pivoted_df.apply(lambda row: fill_exclude_columns(row, pivoted_df.columns), axis=1)
    
    pivoted_df['АФ'] = where(pivoted_df['Регион проведения акции']=='Алматы', 'X', '')
    pivoted_df['АСТ'] = where(pivoted_df['Регион проведения акции']=='Астана', 'X', '')
    pivoted_df['КФ'] = where(pivoted_df['Регион проведения акции']=='Караганда', 'X', '')
    pivoted_df['ШФ'] = where(pivoted_df['Регион проведения акции']=='Шымкент', 'X', '')
    pivoted_df['ТКФ'] = where(pivoted_df['Регион проведения акции']=='Талдыкорган', 'X', '')
    pivoted_df['ППФ'] = where(pivoted_df['Регион проведения акции']=='Петропавловск', 'X', '')
    pivoted_df['КЗФ'] = where(pivoted_df['Регион проведения акции']=='Кызылорда', 'X', '')
    pivoted_df['ТЗФ'] = where(pivoted_df['Регион проведения акции']=='Тараз', 'X', '')
    pivoted_df['ТФ'] = where(pivoted_df['Регион проведения акции']=='Туркестан', 'X', '')
    pivoted_df['УКФ'] = where(pivoted_df['Регион проведения акции']=='Усть-Каменогорск', 'X', '')
    
    pivoted_df = pivoted_df.sort_values(['Отдел', 'Группа', 'Подгруппа', 'Регион проведения акции']).reset_index(drop=True)

    # отсортировать значения в ТФ
    pivoted_df['Исключить ТК Туркестан'] = pivoted_df['Исключить ТК Туркестан'].apply(lambda x: ', '.join(sorted(x.split(', '))) if len(x) > 0 else x)

    """
    Если в регионе всё == Не участвует, то пометить эти строки
    """
    mask_df = pd.DataFrame({'mask': [True] * len(pivoted_df)})
    pivoted_df = pd.concat([pivoted_df, mask_df], axis=1)

    mark_dict = {'АФ': 'Алматы',
                'ФКС': 'Алматы',
                'ЕКФ': 'Алматы',
                'КПФ': 'Алматы',
                'АСТ': 'Астана',
                'КФ': 'Караганда',
                'ШФ': 'Шымкент',
                'ТКФ': 'Талдыкорган',
                'ППФ': 'Петропавловск',
                'КЗФ': 'Кызылорда',
                'ТЗФ': 'Тараз',
                'ТФ': 'Туркестан',
                'УКФ': 'Усть-Каменогорск'}

    for i, row in pivoted_df.iterrows():
        region = row['Регион проведения акции']
        for col in pivoted_df.columns:
            if '\n' in col and col.split('\n')[1] in mark_dict and mark_dict[col.split('\n')[1]] == region:
                if row[col] != 'НЕ УЧАСТВУЕТ':
                    pivoted_df.at[i, 'mask'] = False
                    break
    
    pivoted_df = pivoted_df.loc[pivoted_df['mask']==False].drop(columns=['mask']).reset_index(drop=True)

    def count_values(pivoted_df, value):
        """
        Разметить проблемные модули
        """
        return pivoted_df.groupby(['Регион проведения акции', 'Торговая марка',
            'Регулярная розничная цена с НДС,тг', 'Акционная цена с НДС, тг'])['Модуль'].transform(lambda x: x.eq(str(value)).sum())

    pivoted_df['Единицы'] = count_values(pivoted_df, "1")
    pivoted_df['Нули'] = count_values(pivoted_df, "0")

    def check_problematic_modules(row):
        if row['Единицы'] > 1 or (row['Нули'] >= 1 and row['Единицы'] == 0):
            return 1 #'Не ок'
        else:
            return 0 #'Ок'

    # Колонка потери в файле заполняется формулой. Чтоб не плодить лишние колонки, временно заткнул ее проблемными модулями
    pivoted_df['Скидка для покупателя, %'] = pivoted_df.apply(check_problematic_modules, axis=1)

    pivoted_df = pivoted_df.drop(columns=['Единицы', 'Нули'])

    # to binary flow
    output = io.BytesIO()
    sexy_xlsx(pivoted_df, output)



    return io.BytesIO(output.getvalue())

#%%
